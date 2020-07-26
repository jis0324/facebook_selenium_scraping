# -*- coding: utf-8 -*-
import os
import time
import json
from selenium import webdriver
import traceback
import openpyxl

base_dir = os.path.dirname(os.path.abspath(__file__))

def set_driver():
    user_agent = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_3) AppleWebKit/537.36 (KHTML, like Gecko) ' \
        'Chrome/80.0.3987.132 Safari/537.36'
    chrome_option = webdriver.ChromeOptions()
    chrome_option.add_argument('--no-sandbox')
    chrome_option.add_argument('--disable-dev-shm-usage')
    chrome_option.add_argument('--ignore-certificate-errors')
    chrome_option.add_argument("--disable-blink-features=AutomationControlled")
    chrome_option.add_argument(f'user-agent={user_agent}')
    # chrome_option.headless = True
    
    driver = webdriver.Chrome(options = chrome_option)
    return driver
        
def main():
    url = 'https://www.facebook.com/'
    email = 'working.user101@gmail.com'
    password = 'workinguser01'

    # /* Move to facebook.com */
    driver = set_driver()
    driver.get(url)
    time.sleep(10)
        
    try:
        # /* Login */
        login_email = driver.find_element_by_css_selector("input.login_form_input_box[type='email']")
        login_email.send_keys(email)
        time.sleep(5)

        login_password = driver.find_element_by_css_selector("input.login_form_input_box[type='password']")
        login_password.send_keys(password)
        time.sleep(5)

        driver.find_element_by_css_selector("input[data-testid='royal_login_button']").click()
        time.sleep(10)

        # /* Get Group URL List */
        group_url_list = list()
        with open (base_dir + '/group_urls.txt', 'r') as group_urls_txt:
            group_url_list = [row.strip() for row in group_urls_txt.read().split('\n')]
        
        # /* Loop and Get Group Data */
        for index, group_url in enumerate(group_url_list):
            print('---------------------------------------------')
            print('Getting Data From ', str(index + 1), 'th Group...')
            try:
                driver.get(group_url)
                time.sleep(10)

                result_dict = {
                    "group_name" : "",
                    "group_members" : "",
                    "group_rules" : list(),
                    "group_posts" : list(),
                }

                try:
                    group_name = driver.find_element_by_xpath("//h1[@id='seo_h1_tag']/a")
                    if group_name:
                        result_dict["group_name"] = group_name.text
                        print("group name :", result_dict["group_name"])
                except Exception as err:
                    print(err)
                    pass

                try:
                    group_about_tab = driver.find_element_by_xpath("//a[@title='About']")
                    if group_about_tab:
                        # in the case of public group
                        group_about_tab.click()
                        time.sleep(10)
                except Exception as err:
                    # print(err)
                    pass
                
                try:
                    group_member_count = driver.find_element_by_xpath("//div[@id='pagelet_group_about']/div[2]/div[@role='heading']")
                    if group_member_count:
                        result_dict["group_members"] = group_member_count.text.split("·")[1].strip()
                        print("group members :", result_dict["group_members"])
                except Exception as err:
                    print(err)
                    pass

                try:
                    group_rules_section_heading = driver.find_element_by_xpath("//div[@id='pagelet_group_about']/div[4]/div[@role='heading']")
                    if "Rules" in group_rules_section_heading.text:
                        group_rule_divs = driver.find_elements_by_xpath("//div[@id='pagelet_group_about']/div[4]/div[2]/div/div")
                        group_rules_list = list()
                        for group_rule_div in group_rule_divs:
                            group_rules_list.append(group_rule_div.text.rsplit('\n', 1)[0].replace('\n', '. '))
                        result_dict["group_rules"] = group_rules_list
                    else:
                        result_dict["group_rules"] = ['No Rule']

                except Exception as err:
                    print(err)
                    result_dict["group_rules"] = ['No Rule']
                    pass

                try:
                    group_posts_section_heading = driver.find_element_by_xpath("//div[@id='pagelet_group_about']/div[4]/div[1]")
                    if "Posts" in group_posts_section_heading.text:
                        group_post_divs = driver.find_elements_by_xpath("//div[@id='pagelet_group_about']/div[4]/div")
                        group_posts_list = list()
                            
                        if len(group_post_divs) > 2:
                            for index, group_post_div in enumerate(group_post_divs[1 : len(group_post_divs) - 1]):
                                group_post_dict = {
                                    "post_title" : "",
                                    "post_likes" : "",
                                    "post_comments" : ""
                                }
                                try:
                                    title_div = group_post_div.find_element_by_xpath("//div[@id='pagelet_group_about']/div[4]/div[{}]/div[2]".format(index+2))
                                    if title_div:
                                        group_post_dict['post_title'] = title_div.text
                                except:
                                    pass

                                try:
                                    like_div = group_post_div.find_element_by_xpath("//div[@id='pagelet_group_about']/div[4]/div[{}]/div[3]//div[@class='_1vaq']".format(index+2))
                                    if like_div:
                                        group_post_dict['post_likes'] = like_div.text.strip().split('\n')[0]
                                except:
                                    pass

                                try:
                                    comment_div = group_post_div.find_element_by_xpath("//div[@id='pagelet_group_about']/div[4]/div[{}]/div[3]//div[@class='_ipo']".format(index+2))
                                    if comment_div:
                                        group_post_dict['post_comments'] = comment_div.text.split('Comments')[0].strip()
                                except:
                                    pass

                                group_posts_list.append(group_post_dict)

                                if len(group_posts_list) < 5:
                                    continue
                                else:
                                    break

                        result_dict["group_posts"] = group_posts_list
                        print("group posts : ", result_dict["group_posts"])
                except Exception as err:
                    print(err)
                    pass
                
                try:
                    if not result_dict["group_posts"]:
                        group_posts_section_heading = driver.find_element_by_xpath("//div[@id='pagelet_group_about']/div[5]/div[1]")
                        if "Posts" in group_posts_section_heading.text:
                            group_post_divs = driver.find_elements_by_xpath("//div[@id='pagelet_group_about']/div[5]/div")
                            group_posts_list = list()

                            if len(group_post_divs) > 2:
                                for index, group_post_div in enumerate(group_post_divs[1 : len(group_post_divs) - 1]):
                                    group_post_dict = {
                                        "post_title" : "",
                                        "post_likes" : "",
                                        "post_comments" : ""
                                    }
                                    
                                    try:
                                        title_div = group_post_div.find_element_by_xpath("//div[@id='pagelet_group_about']/div[5]/div[{}]/div[2]".format(index+2))
                                        if title_div:
                                            group_post_dict['post_title'] = title_div.text
                                    except:
                                        pass

                                    try:
                                        like_div = group_post_div.find_element_by_xpath("//div[@id='pagelet_group_about']/div[5]/div[{}]/div[3]//div[@class='_1vaq']".format(index+2))
                                        if like_div:
                                            group_post_dict['post_likes'] = like_div.text.strip().split('\n')[0]
                                    except:
                                        pass

                                    try:
                                        comment_div = group_post_div.find_element_by_xpath("//div[@id='pagelet_group_about']/div[5]/div[{}]/div[3]//div[@class='_ipo']".format(index+2))
                                        if comment_div:
                                            group_post_dict['post_comments'] = comment_div.text.split('Comments')[0].strip()
                                    except:
                                        pass

                                    group_posts_list.append(group_post_dict)

                                    if len(group_posts_list) < 5:
                                        continue
                                    else:
                                        break

                            result_dict["group_posts"] = group_posts_list
                            print("group posts : ", result_dict["group_posts"])
                            
                except Exception as err:
                    # print(err)
                    pass
                
                while len(result_dict["group_posts"]) < 5:
                    result_dict["group_posts"].append({
                        "post_title" : "",
                        "post_likes" : "",
                        "post_comments" : ""
                    })
                
                print('-----------------------------------------')
                print(result_dict)
                
                result_excel_file = openpyxl.load_workbook('group_data.xlsx')
                allSheetNames = result_excel_file.sheetnames

                currentSheet = result_excel_file['Sheet1']

                max_row = currentSheet.max_row

                for i in range(5):
                    currentSheet.row_dimensions[i + 1 + max_row].height = 40

                currentSheet.merge_cells('A{}:A{}'.format(max_row + 1, max_row + 5))
                cell_name = "{}{}".format('A', max_row + 1)
                currentSheet[cell_name] =  ((max_row - 1) / 5) + 1
                currentSheet[cell_name].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True) 

                currentSheet.merge_cells('B{}:B{}'.format(max_row + 1, max_row + 5))
                cell_name = "{}{}".format('B', max_row + 1)
                currentSheet[cell_name] = result_dict["group_name"]
                currentSheet[cell_name].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True) 

                currentSheet.merge_cells('C{}:C{}'.format(max_row + 1, max_row + 5))
                cell_name = "{}{}".format('C', max_row + 1)
                currentSheet[cell_name] = result_dict["group_members"]
                currentSheet[cell_name].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True) 

                currentSheet.merge_cells('D{}:D{}'.format(max_row + 1, max_row + 5))
                cell_name = "{}{}".format('D', max_row + 1)
                group_rules_text = ''
                if result_dict["group_rules"]:
                    for row in result_dict["group_rules"]:
                        group_rules_text += row + '\n'
                else:
                    group_rules_text = 'No Rule'
                currentSheet[cell_name] = group_rules_text
                currentSheet[cell_name].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True) 

                for i in range(5):
                    # post title
                    cell_name = "{}{}".format('E', i + 1 + max_row)
                    currentSheet[cell_name] = result_dict["group_posts"][i]['post_title']
                    currentSheet[cell_name].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

                    # post likes
                    cell_name = "{}{}".format('F', i + 1 + max_row)
                    currentSheet[cell_name] = result_dict["group_posts"][i]['post_likes']
                    currentSheet[cell_name].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

                    # post comments
                    cell_name = "{}{}".format('G', i + 1 + max_row)
                    currentSheet[cell_name] = result_dict["group_posts"][i]['post_comments']
                    currentSheet[cell_name].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True) 

                result_excel_file.save(filename="group_data.xlsx")
            except:
                continue

        driver.quit()
    except Exception as err :
        print(err)
        driver.quit()
        return

if __name__ == '__main__':
    main()
